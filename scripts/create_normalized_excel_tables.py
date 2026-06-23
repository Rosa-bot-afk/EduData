from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
INPUT = ROOT / "data" / "official_students_longitudinal_90000_v2.xlsx"
OUTPUT = ROOT / "data" / "normalized_sql_tables.xlsx"

EXCEL_MAX_ROWS = 1_048_576
DATA_ROWS_PER_SHEET = EXCEL_MAX_ROWS - 1

METRICS = [
    ("family_income_level", "socioeconomic", "coded_category"),
    ("internet_access_hours", "digital_access", "hours_per_day"),
    ("academic_motivation", "academic", "score_1_10"),
    ("social_media_hours", "digital_behavior", "hours_per_day"),
    ("sessions_per_day", "digital_behavior", "count_per_day"),
    ("average_session_length_minutes", "digital_behavior", "minutes"),
    ("late_night_usage", "digital_behavior", "coded_category"),
    ("education_content_hours", "content", "hours_per_day"),
    ("short_video_hours", "content", "hours_per_day"),
    ("entertainment_content_hours", "content", "hours_per_day"),
    ("news_content_hours", "content", "hours_per_day"),
    ("likes_given_per_day", "interaction", "count_per_day"),
    ("comments_written_per_day", "interaction", "count_per_day"),
    ("posts_created_per_week", "interaction", "count_per_week"),
    ("brain_rot_level", "wellbeing", "percent"),
    ("attention_span_minutes", "academic", "minutes"),
    ("study_hours_per_week", "academic", "hours_per_week"),
    ("class_attendance_rate", "academic", "percent"),
    ("productivity_score", "wellbeing", "score_0_10"),
    ("sleep_hours", "health", "hours_per_night"),
    ("stress_level", "health", "score_0_10"),
    ("anxiety_score", "health", "score_0_10"),
    ("depression_score", "health", "score_0_10"),
    ("ads_viewed_per_day", "advertising", "count_per_day"),
    ("ads_clicked_per_week", "advertising", "count_per_week"),
    ("impulse_purchase_scale", "spending", "score_0_10"),
    ("digital_spending_per_month", "spending", "money_per_month"),
    ("cyberbullying_exposure", "risk", "coded_boolean"),
    ("adult_content_exposure", "risk", "coded_boolean"),
    ("digital_addiction_score", "wellbeing", "score_0_100"),
    ("wellbeing_percent", "wellbeing", "percent"),
    ("perceived_financial_impact", "spending", "percent"),
]

CODE_MAPS = {
    "family_income_level": {"LOW": 1, "MIDDLE": 2, "HIGH": 3},
    "late_night_usage": {"Never": 0, "Sometimes": 1, "Often": 2, "Always": 3},
    "cyberbullying_exposure": {"No": 0, "Yes": 1},
    "adult_content_exposure": {"No": 0, "Yes": 1},
}


def append_dataframe(ws, df: pd.DataFrame) -> None:
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))


def coded_value(metric_name: str, raw_value):
    if metric_name in CODE_MAPS:
        return CODE_MAPS[metric_name][str(raw_value).strip()]
    return float(raw_value)


def main() -> None:
    print("Reading source workbook...")
    data = pd.read_excel(INPUT, sheet_name="DATA")
    users = pd.read_excel(INPUT, sheet_name="USERS")
    employees = pd.read_excel(INPUT, sheet_name="EMPLOYEES")

    data = data.sort_values(["student_id", "survey_number"]).reset_index(drop=True)
    student_data = data.drop_duplicates("student_id").copy()

    development = (
        pd.DataFrame({"Development_level_name": sorted(data["development_level"].unique())})
        .reset_index()
        .rename(columns={"index": "Development_Level_id"})
    )
    development["Development_Level_id"] += 1
    development_id = dict(zip(development["Development_level_name"], development["Development_Level_id"]))

    country = (
        student_data[["Country", "development_level"]]
        .drop_duplicates()
        .sort_values("Country")
        .reset_index(drop=True)
    )
    country.insert(0, "country_id", range(1, len(country) + 1))
    country["development_id"] = country["development_level"].map(development_id)
    country = country[["country_id", "Country", "development_id"]].rename(columns={"Country": "country_name"})
    country_id = dict(zip(country["country_name"], country["country_id"]))

    country_indicator = (
        student_data[
            [
                "Country",
                "poverty_rate_percent",
                "internet_infrastructure_index",
                "average_internet_speed_mbps",
            ]
        ]
        .drop_duplicates("Country")
        .sort_values("Country")
        .reset_index(drop=True)
    )
    country_indicator.insert(0, "Country_Indicator_id", range(1, len(country_indicator) + 1))
    country_indicator["indicator_year"] = 2025
    country_indicator["country_id"] = country_indicator["Country"].map(country_id)
    country_indicator = country_indicator[
        [
            "Country_Indicator_id",
            "indicator_year",
            "poverty_rate_percent",
            "internet_infrastructure_index",
            "average_internet_speed_mbps",
            "country_id",
        ]
    ].rename(columns={"average_internet_speed_mbps": "average_internet_speed_mbs"})

    university = (
        student_data[["Country", "University"]]
        .drop_duplicates()
        .sort_values(["Country", "University"])
        .reset_index(drop=True)
    )
    university.insert(0, "university_id", range(1, len(university) + 1))
    university["university_type"] = "University"
    university["country_id"] = university["Country"].map(country_id)
    university = university[["university_id", "University", "university_type", "country_id"]].rename(columns={"University": "university_name"})
    university_id = {(row.Country, row.University): row.university_id for row in student_data[["Country", "University"]].drop_duplicates().merge(university, left_on="University", right_on="university_name").itertuples()}

    device = pd.DataFrame({"Device_name": sorted(data["device_access"].unique())})
    device.insert(0, "Device_type_id", range(1, len(device) + 1))
    device_id = dict(zip(device["Device_name"], device["Device_type_id"]))

    academic_period = (
        data[["AcademicPeriod", "start_date", "end_date"]]
        .drop_duplicates()
        .sort_values("AcademicPeriod")
        .reset_index(drop=True)
    )
    academic_period.insert(0, "Academic_Period_id", range(1, len(academic_period) + 1))
    academic_period["Period_year"] = academic_period["AcademicPeriod"].astype(str).str.split("-").str[0].astype(int)
    academic_period["period_term"] = academic_period["AcademicPeriod"].astype(str).str.split("-").str[1].astype(int)
    academic_period = academic_period[["Academic_Period_id", "start_date", "Period_year", "period_term", "end_date"]]
    academic_period_id = {(row.Period_year, row.period_term): row.Academic_Period_id for row in academic_period.itertuples()}

    field = pd.DataFrame({"field_of_study": sorted(data["field_of_study"].unique())})
    field.insert(0, "field_of_study_id", range(1, len(field) + 1))
    field_id = dict(zip(field["field_of_study"], field["field_of_study_id"]))

    metric_definition = pd.DataFrame(
        [
            {"Metric_id": i, "metric_name": name, "category": category, "unit": unit}
            for i, (name, category, unit) in enumerate(METRICS, start=1)
        ]
    )
    metric_id = dict(zip(metric_definition["metric_name"], metric_definition["Metric_id"]))

    user_table = users[["usuario_id", "Email", "User", "Password", "Created_at", "Rol"]].rename(
        columns={
            "usuario_id": "user_id",
            "Email": "email",
            "User": "user_name",
            "Password": "password",
            "Rol": "user_role",
        }
    )

    student = student_data[["student_id", "Gender", "Country", "University", "usuario_id"]].copy()
    student["university_id"] = student.apply(lambda row: university_id[(row["Country"], row["University"])], axis=1)
    student["country_id"] = student["Country"].map(country_id)
    student = student[["student_id", "Gender", "university_id", "country_id", "usuario_id"]].rename(
        columns={"Gender": "gender"}
    )

    employee_user = users.dropna(subset=["employee_id"]).copy()
    employee_user["employee_id"] = employee_user["employee_id"].astype(int)
    employee_user_id = dict(zip(employee_user["employee_id"], employee_user["usuario_id"]))
    employee = employees.copy()
    employee["user_id"] = employee["employee_id"].map(employee_user_id)
    employee = employee[["employee_id", "First_name", "LastName", "hire_date", "job_rol", "user_id"]].rename(columns={"job_rol": "job_role"})

    first_assessment = data.drop_duplicates("student_id").copy()
    adult = first_assessment[first_assessment["Age"] >= 18][["autonomia_financiera", "trabaja_actualmente", "student_id"]].copy()
    adult.insert(0, "adult_student_id", range(1, len(adult) + 1))
    adult = adult.rename(columns={"autonomia_financiera": "financial_autonomy", "trabaja_actualmente": "is_employed"})

    minor = first_assessment[first_assessment["Age"] < 18][["tipo_tutor", "consentimiento_tutor", "student_id"]].copy()
    minor.insert(0, "minor_student_id", range(1, len(minor) + 1))
    minor = minor.rename(columns={"tipo_tutor": "guardian_type", "consentimiento_tutor": "guardian_consent"})

    nationality = pd.DataFrame({"nacionality_name": sorted(student_data["Nationality"].unique())})
    nationality.insert(0, "nacionality_id", range(1, len(nationality) + 1))
    nationality_id = dict(zip(nationality["nacionality_name"], nationality["nacionality_id"]))
    student_nationality = student_data[["student_id", "Nationality"]].copy()
    student_nationality["nacionality_id"] = student_nationality["Nationality"].map(nationality_id)
    student_nationality = student_nationality[["nacionality_id", "student_id"]]

    assessment = data[[
        "student_id",
        "survey_number",
        "urban_rural",
        "education_level",
        "sent_date",
        "reception_date",
        "parents_divorced",
        "Age",
        "device_access",
        "AcademicPeriod",
        "field_of_study",
    ]].copy()
    assessment.insert(0, "Assessment_id", range(1, len(assessment) + 1))
    assessment["Device_type_id"] = assessment["device_access"].map(device_id)
    assessment["period_year"] = assessment["AcademicPeriod"].astype(str).str.split("-").str[0].astype(int)
    assessment["period_term"] = assessment["AcademicPeriod"].astype(str).str.split("-").str[1].astype(int)
    assessment["academic_period_id"] = assessment.apply(lambda row: academic_period_id[(row["period_year"], row["period_term"])], axis=1)
    assessment["field_of_study_id"] = assessment["field_of_study"].map(field_id)
    assessment = assessment[[
        "Assessment_id",
        "survey_number",
        "urban_rural",
        "education_level",
        "sent_date",
        "reception_date",
        "parents_divorced",
        "Age",
        "Device_type_id",
        "academic_period_id",
        "field_of_study_id",
        "student_id",
    ]].rename(columns={"Age": "age"})

    print("Writing normalized workbook...")
    wb = Workbook(write_only=True)
    for sheet_name, df in [
        ("DevelopmentLevel", development),
        ("Country", country),
        ("CountryIndicator", country_indicator),
        ("University", university),
        ("DeviceType", device),
        ("AcademicPeriod", academic_period),
        ("FieldOfStudy", field),
        ("MetricDefinition", metric_definition),
        ("User", user_table),
        ("Student", student),
        ("Employee", employee),
        ("AdultStudent", adult),
        ("MinorStudent", minor),
        ("Nacionality", nationality),
        ("Student_nacionality", student_nationality),
        ("StudentAssessment", assessment),
    ]:
        ws = wb.create_sheet(sheet_name)
        append_dataframe(ws, df)

    sheet_number = 1
    sheet_row_count = 0
    ws_mide = wb.create_sheet(f"Mide_{sheet_number}")
    ws_mide.append(["metric_id", "Assessment_id", "metric_value"])
    for row in data.itertuples(index=False):
        row_dict = row._asdict()
        assessment_id_value = int(row_dict["Index"]) + 1 if "Index" in row_dict else None
        if assessment_id_value is None:
            assessment_id_value = int(row_dict["student_id"] - 1) * 6 + int(row_dict["survey_number"])
        for metric_name, _, _ in METRICS:
            if sheet_row_count >= DATA_ROWS_PER_SHEET:
                sheet_number += 1
                sheet_row_count = 0
                ws_mide = wb.create_sheet(f"Mide_{sheet_number}")
                ws_mide.append(["metric_id", "Assessment_id", "metric_value"])
            ws_mide.append([metric_id[metric_name], assessment_id_value, coded_value(metric_name, row_dict[metric_name])])
            sheet_row_count += 1

    wb.save(OUTPUT)
    print(OUTPUT)
    print("Done.")


if __name__ == "__main__":
    main()
